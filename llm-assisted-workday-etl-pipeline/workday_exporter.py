"""
workday_exporter.py
-------------------
Stage 5: Generate the Workday EIB (Enterprise Interface Builder) export
from the mapped and reviewed rows in SQLite.

Per the process flow doc Stage 5:
  - Generate Workday EIB (Excel template) or call Workday SOAP/REST API.
  - Run validation against the Workday job catalog before submitting.
  - Every load produces a workday_load_log row with the submission ID.

Design choices:
  - Only rows with review_status IN ("auto", "approved") are exported.
    "review" rows (confidence < 0.8) are held back until a human signs off.
    "overridden" rows are included -- reviewer made a deliberate correction.
  - Output is a dated Excel file matching Workday EIB column expectations.
  - A workday_load_log table is written back to SQLite for audit trail.
  - Pre-export validation checks every job code and location code against
    the known catalogs before the file is generated.
"""

import pandas as pd
from sqlalchemy import create_engine
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os


# =========================================================
# WORKDAY EIB COLUMN MAPPING
# Maps our SQLite column names to Workday EIB header names.
# Extend this when Workday adds/renames fields.
# =========================================================

EIB_COLUMN_MAP = {
    "source_record_id":    "External_ID",
    "First_Name":          "First_Name",
    "Last_Name":           "Last_Name",
    "Birth_Date":          "Date_of_Birth",
    "Hire_Date":           "Hire_Date",
    "Gender":              "Gender",
    "Email":               "Work_Email",
    "workday_job_code":    "Job_Profile",
    "workday_department":  "Supervisory_Org",
    "workday_location_code": "Location",
    "Pay_Type":            "Pay_Rate_Type",
    "Annual_Salary":       "Annual_Base_Pay",
}

# Catalog sets for pre-export validation
VALID_JOB_CODES = {
    "PHYS_FAM", "PHYS_INTERN", "NP_FAM", "RN", "LPN", "MA",
    "PHARM_RX", "PHARM_TECH", "DENT_DDS", "DENT_HYG", "BH_LCSW",
    "ADMIN_FD", "ADMIN_PM", "ADMIN_CMO", "BILL_SPEC", "BILL_CODER",
    "IT_SUP", "HR_SPEC", "FIN_ANALYST", "OPS_MGR",
}

VALID_LOCATION_CODES = {
    "AUG", "BAT", "SRCY", "HBR", "MTN",
    "NWP", "CBT", "BEE", "CLN", "CRN", "REM",
}


def load_export_ready_rows(db_path: str = "sqlite:///employees.db") -> pd.DataFrame:
    """
    Loads rows that are cleared for Workday export.
    Excludes "review" status -- those need human sign-off first.
    """
    engine = create_engine(db_path)
    df = pd.read_sql(
        """
        SELECT * FROM workday_employees
        WHERE review_status IN ('auto', 'approved', 'overridden')
        """,
        engine,
    )
    return df, engine


def validate_before_export(df: pd.DataFrame) -> list[str]:
    """
    Runs pre-export checks against Workday catalogs.
    Returns a list of validation error strings (empty = all clear).
    Per process flow doc: 'Run validation against the Workday job
    catalog before submitting.'
    """
    errors = []

    # Job code check
    invalid_jobs = df[
        ~df["workday_job_code"].isin(VALID_JOB_CODES) &
        df["workday_job_code"].notnull()
    ]
    if len(invalid_jobs):
        for _, row in invalid_jobs.iterrows():
            errors.append(
                f"Row {row['source_record_id']}: invalid job code "
                f"'{row['workday_job_code']}'"
            )

    # Location code check
    invalid_locs = df[
        ~df["workday_location_code"].isin(VALID_LOCATION_CODES) &
        df["workday_location_code"].notnull()
    ]
    if len(invalid_locs):
        for _, row in invalid_locs.iterrows():
            errors.append(
                f"Row {row['source_record_id']}: invalid location code "
                f"'{row['workday_location_code']}'"
            )

    # Missing required fields
    for col in ["First_Name", "Last_Name", "Hire_Date", "workday_job_code"]:
        missing = df[df[col].isnull()]
        if len(missing):
            errors.append(
                f"{len(missing)} rows missing required field '{col}'"
            )

    return errors


def build_eib_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Selects and renames columns to match Workday EIB headers.
    Only exports columns Workday expects -- audit/staging columns
    are intentionally excluded from the EIB file.
    """
    available = {k: v for k, v in EIB_COLUMN_MAP.items() if k in df.columns}
    eib_df = df[list(available.keys())].copy()
    eib_df = eib_df.rename(columns=available)

    # ISO-format dates for Workday
    for date_col in ["Date_of_Birth", "Hire_Date"]:
        if date_col in eib_df.columns:
            eib_df[date_col] = pd.to_datetime(
                eib_df[date_col], errors="coerce"
            ).dt.strftime("%Y-%m-%d")

    return eib_df


def write_eib_excel(eib_df: pd.DataFrame, output_path: str) -> None:
    """
    Writes the EIB DataFrame to a formatted Excel file.
    Uses openpyxl directly for header styling -- Workday EIB
    files are reviewed by humans before upload so readability matters.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Worker_EIB"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(eib_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(eib_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_path)


def log_load(engine, row_count: int, output_path: str, errors: list) -> None:
    """
    Writes a workday_load_log entry per process flow doc Stage 5:
    'Every load produces a workday_load_log row with the submission ID.'
    Submission ID is the filename timestamp here; replace with Workday
    API response ID once the API integration is live.
    """
    log_df = pd.DataFrame([{
        "submission_id":   os.path.basename(output_path),
        "exported_rows":   row_count,
        "export_timestamp": datetime.now().isoformat(),
        "validation_errors": "; ".join(errors) if errors else "none",
        "status":          "exported" if not errors else "exported_with_warnings",
    }])
    log_df.to_sql("workday_load_log", engine, if_exists="append", index=False)


def run_export(db_path: str = "sqlite:///employees.db") -> str:
    """
    Full Stage 5 export. Returns the output file path.
    """
    print("\n[Stage 5] Workday EIB Export")

    df, engine = load_export_ready_rows(db_path)
    held_back = pd.read_sql(
        "SELECT COUNT(*) as n FROM workday_employees WHERE review_status = 'review'",
        engine,
    ).iloc[0]["n"]

    print(f"  Export-ready rows:  {len(df)}")
    print(f"  Held for review:    {held_back}  (run dashboard to approve)")

    if len(df) == 0:
        print("  Nothing to export.")
        return None

    # Validate before generating the file
    errors = validate_before_export(df)
    if errors:
        print(f"  ⚠ {len(errors)} validation warning(s):")
        for e in errors:
            print(f"    - {e}")
    else:
        print("  Validation passed -- all job codes and location codes valid.")

    eib_df = build_eib_dataframe(df)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"workday_eib_{timestamp}.xlsx"
    write_eib_excel(eib_df, output_path)

    log_load(engine, len(eib_df), output_path, errors)

    print(f"  EIB file written:   {output_path}")
    print(f"  Load log updated in employees.db (workday_load_log)")

    return output_path


if __name__ == "__main__":
    run_export()